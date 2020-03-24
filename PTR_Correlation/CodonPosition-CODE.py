
############################################################################################
# CALCULATE THE POSITION OF THE SPECIFIED CODON IN THE SEQUENCES
############################################################################################

#import relevant libraries
import pandas as pd
import openpyxl as xl 

# opening the source excel file 
wb1 = xl.load_workbook('Table_EV4.xlsx')
ws1 = wb1.worksheets[0] 
  
# opening the destination excel file  
# Create a Workbook
wb2 = xl.Workbook()
ws2 =  wb2.active
ws2.title = "Positions"
wb2.save(filename = 'Pos.xlsx')

wb2 = xl.load_workbook('Pos.xlsx') 
ws2 = wb2.active
  
# calculate total number of rows and  
# columns in source excel file 
mr = ws1.max_row 
mc = ws1.max_column 
  
# copying the cell values from source  
# excel file to destination excel file 
for i in range (1, mr + 1): 
    for j in range (1, mc + 1): 
        # reading cell value from source excel file 
        c = ws1.cell(row = i, column = j) 
  
        # writing the read value to destination excel file 
        ws2.cell(row = i, column = j).value = c.value 
  
# saving the destination excel file 
wb2.save(str('Pos.xlsx'))

#code for CodonPosition function 
from Bio.Seq import Seq
from Bio.Alphabet import IUPAC
my_seq = "GAUCGAUGGGCUUAUAUAGGAUCGAAAAUCGCA"

aa = 'AAG' # select codon to find data for
def CodonPosition (my_seq):
    codons = []
    start = 0
    # separate the sequences into a codons list
    for l in range(int(len(my_seq)/3)):
        end = start + 3
        codon = my_seq[start:end]
        codons.append(codon)
        start = end
    #print(codons)
    
    # Find only the position of the selected codon
    if aa not in codons:
        pos = '00' #if not in sequence, output '00'
    else:
        pos = codons.index(aa) #if in sequence, output the position on the list

    PosDict = {aa: pos} #create a dictionary for the codon and its positions
    
    return PosDict

#print(CodonPosition(my_seq))

xls_file = pd.ExcelFile('Pos.xlsx') # Import the excel file and call it xls_file
df = xls_file.parse() #import into pandas dataframe object  

db_len = len(df)

gene_ids = df.get('CDS_Sequence')#only the CDS_Sequence column
output = pd.DataFrame()#empt7 df
for i in range(len(gene_ids)):
    seq = gene_ids[i]
    codon = CodonPosition(seq)#use of CodontTable fonction that outputs a dictionary 
    output = output.append(codon, ignore_index=True)#append as rows with column the dictionary keys
#print(output.head())

df_merge_col = pd.merge(df, output, left_index=True, right_index=True)# use merge method, not join to add with respect to rows
df_merge_col.to_excel('Pos.xlsx',index=False)#use this method to save to csv, traditinal format, better than excel


###########################################################################################################
## Make column of PTR for SPECIFIC TISSUE (BRAIN)
###########################################################################################################

import pandas as pd
import json 
import openpyxl as xl 
from openpyxl import load_workbook
from decimal import Decimal
import math

wb = load_workbook('Pos.xlsx')

ws1 = wb.worksheets[0]
ws1.title = "Positions" 
  
# opening the source excel file 
wbEV3 = xl.load_workbook('Table_EV3.xlsx')
wsEV3 = wbEV3.worksheets[0] 

# calculate total number of rows and  
# columns in source excel file 
mrEV3 = wsEV3.max_row 
mcEV3 = wsEV3.max_column 

mr = ws1.max_row 
mc = ws1.max_column
    
# copying the cell values from source  
# excel file to destination excel file 
for j in range (1, mcEV3 + 1):
    for i in range (1, mrEV3 + 1): 
        # reading cell value from source excel file 
        c = wsEV3.cell(row = i, column = j)
        h = wsEV3.cell(row = 1, column = j)

        c_val = c.value
        h_val = h.value

        if (h_val == 'Brain_PTR'):
            ws1.cell(row = i, column = mc+1).value = c.value 

wb.save(filename = 'Pos.xlsx')

####################################################################################################
#Create data for scatterplot
#####################################################################################################

# import pandas
import pandas as pd
# import matplotlib
import matplotlib.pyplot as plt
# import seaborn
import seaborn as sns

df = pd.read_excel('Pos.xlsx', sheet_name='Positions')
print(df)

df = df.drop(columns=['EnsemblGeneID', 'EnsemblTranscriptID', 'EnsemblProteinID', 'Unnamed: 4'])
print(df)

sns.scatterplot(x="AAG", y="Brain_PTR", data=df)

plt.rcParams["figure.figsize"] = (30, 8)
plt.ylabel(("PTR value of Brain Tissue"))
plt.xlabel(("AAG codon initial position"))
plt.xticks(rotation=90)

plt.show()