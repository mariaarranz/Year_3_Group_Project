
########################################################################################################
# Create Workbook and add Codon Frequencies Table
###########################################################################################################

from decimal import Decimal
import math
from openpyxl import Workbook

wb = Workbook()
wb.save(filename='PearsonCorrelation.xlsx')

ws2 = wb.create_sheet("Codons", 0)
ws2.title = "Codon Frequencies"

from openpyxl import load_workbook
wb1 = load_workbook('CodonTableBias.xlsx')
ws1 = wb1.worksheets[0]

# calculate total number of rows and  
# columns in source excel file 
mr = ws1.max_row 
mc = ws1.max_column 

# copying the cell values from source  
# excel file to destination excel file 
for i in range (1, mr + 1): 
    for j in range (1, mc + 1): 
        # reading cell value from source excel file 
        c = ws1.cell(row = i, column = j)
        h = ws1.cell(row = 1, column = j)

        c_val = c.value
        h_val = h.value

        ws2.cell(row = i, column = j).value = c.value 

wb.save(filename = 'PearsonCorrelation.xlsx')

###########################################################################################################
## MAKE NEW SHEET WITH PTR values FOR EACH GENE
###########################################################################################################

import pandas as pd
import json 
import openpyxl as xl 
from openpyxl import load_workbook
from decimal import Decimal
import math

wb = load_workbook('PearsonCorrelation.xlsx')

ws1 = wb.worksheets[0]

ws3 = wb.create_sheet("Sheet_Y", 1)
ws3.title = "PTR"
  
# opening the source excel file 
wbEV3 = xl.load_workbook('Table_EV3.xlsx')
wsEV3 = wbEV3.worksheets[0] 

# calculate total number of rows and  
# columns in source excel file 
mr = wsEV3.max_row 
mc = wsEV3.max_column 

def getYbar(col):
    ls = []
    for i in range (2, 11577): 
        # reading cell value from source excel file 
        c = wsEV3.cell(row = i, column = col)
        c_val = c.value
    
        ls.append(c_val)
    
    summ = 0
    for n in ls:
        summ  += n

    ybar = float(summ/11574)
    return ybar
    
# copying the cell values from source  
# excel file to destination excel file 
for j in range (1, mc + 1): 
    for i in range (1, mr + 1): 
        # reading cell value from source excel file 
        c = wsEV3.cell(row = i, column = j)
        h = wsEV3.cell(row = 1, column = j)

        c_val = c.value
        h_val = h.value

        if (j < 5):
            ws3.cell(row = i, column = j).value = c.value 

        if (j > 4) and (i == 1):
            ws3.cell(row = i, column = j).value = c.value 

        if (j > 4) and (i!=1):
            y = float(c_val)
            ws3.cell(row = i, column = j).value = y

wb.save(filename = 'PearsonCorrelation.xlsx')

###########################################################################################################
# CREATE BOXPLOT DATA FOR FIGURE
###########################################################################################################

# import pandas
import pandas as pd
# import matplotlib
import matplotlib.pyplot as plt
# import seaborn
import seaborn as sns

df = pd.read_excel('PearsonCorrelation.xlsx', sheet_name='Codon Frequencies')
#print(df)

dfY = pd.read_excel('PearsonCorrelation.xlsx', sheet_name='PTR')
#print(dfY)

dfPTR = dfY['Colon_PTR']
#print(dfPTR)

df['PTR'] = dfPTR
df = df.drop(columns=['EnsemblGeneID', 'EnsemblTranscriptID', 'EnsemblProteinID'])
print(df)


#Reference - https://levelup.gitconnected.com/pearson-coefficient-of-correlation-using-pandas-ca68ce678c04

corr = df.corr(method='pearson')
print(corr)

f = plt.figure(1)
sns.heatmap(corr, 
            xticklabels=corr.columns,
            yticklabels=corr.columns,
            cmap='RdBu_r',
            annot=False,
            linewidth=0.5)


corr = corr.drop(['PTR'])
print(corr)

# list(data)
header = list(corr.columns) 

codons = []
trash = []
for i in range(0,64):
    codons.append(header[i])

corr = corr.drop(columns=codons). assign(Codons = codons)
print(corr)

g = plt.figure(2)
sns.scatterplot(x="Codons", y="PTR", data=corr)

xs = corr.Codons
ys = corr.PTR

for x,y in zip(xs,ys):

    label1 = "{}".format(x)
    label2 = ", {:.2f}".format(y)
    label = label1 + label2

    plt.annotate(label, # this is the text
                 (x,y), # this is the point to label
                 textcoords="offset points", # how to position the text
                 xytext=(0,10), # distance from text to points (x,y)
                 ha='center') # horizontal alignment can be left, right or center


plt.rcParams["figure.figsize"] = (30, 8)
plt.ylabel(("Pearson Correlation Coefficient Value"))
plt.xticks(rotation=90)

plt.show()
