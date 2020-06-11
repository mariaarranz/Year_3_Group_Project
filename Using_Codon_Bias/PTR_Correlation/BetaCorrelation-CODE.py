########################################################################################################
# Create Workbook and add Codon Frequencies Table
###########################################################################################################

from decimal import Decimal
import math
from openpyxl import Workbook

wb = Workbook()
wb.save(filename='Beta_Correlation.xlsx')

ws2 = wb.create_sheet("Codons", 0)
ws2.title = "Codon Frequencies"

from openpyxl import load_workbook
wb1 = load_workbook('CodonTableBias.xlsx')
ws1 = wb1.worksheets[0]

# calculate total number of rows and  
# columns in source excel file 
mr = ws1.max_row 
mc = ws1.max_column 

# copying the cell values from source  
# excel file to destination excel file 
for i in range (1, mr + 1): 
    for j in range (1, mc + 1): 
        # reading cell value from source excel file 
        c = ws1.cell(row = i, column = j)
        h = ws1.cell(row = 1, column = j)

        c_val = c.value
        h_val = h.value

        ws2.cell(row = i, column = j).value = c.value 

wb.save(filename = 'Beta_Correlation.xlsx')
print("File Created")

###########################################################################################################
## MAKE NEW SHEET WITH SEQUENCE PREDICTORS FOR EACH CODON - AVERAGE (Xi - X-BAR)
###########################################################################################################

import pandas as pd
import json 
import openpyxl as xl 
from openpyxl import load_workbook
from decimal import Decimal
import math

wb = load_workbook('Beta_Correlation.xlsx')

ws2 = wb.create_sheet("Sheet_X", 1)
ws2.title = "Xi-Xbar"

ws1 = wb.worksheets[0]
ws1.title = "Codon Frequencies"  

# calculate total number of rows and  
# columns in source excel file 
mr = ws1.max_row 
mc = ws1.max_column 

def getXbar(col):
    ls = []
    for i in range (2, 11574): 
        # reading cell value from source excel file 
        c = ws1.cell(row = i, column = col)
        c_val = c.value

        xi = float(c_val)

        if (xi <= 0):
            x = 0
        else:
            x = math.log2(xi)

        ls.append(x)
    
    summ = 0
    for n in ls:
        summ  += n

    xbar = float(summ/11574)

    return xbar
    
# copying the cell values from source  
# excel file to destination excel file 
for j in range (1, mc + 1): 
    if (j > 4):
        xbar = getXbar(j)
    for i in range (1, mr + 1): 
        # reading cell value from source excel file 
        c = ws1.cell(row = i, column = j)
        h = ws1.cell(row = 1, column = j)

        c_val = c.value
        h_val = h.value

        if (j < 5):
            ws2.cell(row = i, column = j).value = c.value 

        if (j > 4) and (i == 1):
            ws2.cell(row = i, column = j).value = c.value 

        if (j > 4) and (i!=1):
            xi = float(c_val)
            if (xi <= 0):
                x = 0
            else:
                x = math.log2(xi)

            sub = float(x - xbar)
            ws2.cell(row = i, column = j).value = sub

wb.save(filename = 'Beta_Correlation.xlsx')
print("X Created")

###########################################################################################################
## MAKE NEW SHEET WITH PTR values FOR EACH GENE - AVERAGE (Yi - Y-BAR)
###########################################################################################################

import pandas as pd
import json 
import openpyxl as xl 
from openpyxl import load_workbook
from decimal import Decimal
import math

wb = load_workbook('Beta_Correlation.xlsx')

ws3 = wb.create_sheet("Sheet_Y", 2)
ws3.title = "Yi-Ybar"
  
# opening the source excel file 
wbEV3 = xl.load_workbook('Table_EV3.xlsx')
wsEV3 = wbEV3.worksheets[0] 

# calculate total number of rows and  
# columns in source excel file 
mr = wsEV3.max_row 
mc = wsEV3.max_column 

def getYbar(col):
    ls = []
    for i in range (2, 11577): 
        # reading cell value from source excel file 
        c = wsEV3.cell(row = i, column = col)
        c_val = c.value
    
        ls.append(c_val)
    
    summ = 0
    for n in ls:
        summ  += n

    ybar = float(summ/11574)

    return ybar
    
# copying the cell values from source  
# excel file to destination excel file 
for j in range (1, mc + 1): 
    if (j > 4):
        ybar = getYbar(j)
    for i in range (1, mr + 1): 
        # reading cell value from source excel file 
        c = wsEV3.cell(row = i, column = j)
        h = wsEV3.cell(row = 1, column = j)

        c_val = c.value
        h_val = h.value

        if (j < 5):
            ws3.cell(row = i, column = j).value = c.value 

        if (j > 4) and (i == 1):
            ws3.cell(row = i, column = j).value = c.value 

        if (j > 4) and (i!=1):
            yi = float(c_val)
            sub = float(yi - ybar)
            if (sub < 0):
                sub = 0
            ws3.cell(row = i, column = j).value = sub

wb.save(filename = 'Beta_Correlation.xlsx')
print("Y Created")

#############################################################################################################
#ADD PTR-AI SHEET TO EXCEL FILE
#############################################################################################################

from openpyxl import load_workbook
from decimal import Decimal
import math

wb = load_workbook('Beta_Correlation.xlsx')

wsAI = wb.create_sheet("Sheet_PTRAI", 3)
wsAI.title = "PTR-AI"

wsX = wb.worksheets[1]
wsY = wb.worksheets[2]

# calculate total number of rows and  
# columns in source excel file 
mrX = wsX.max_row 
mcX = wsX.max_column 

mrY = wsY.max_row 
mcY = wsY.max_column 

def getYvalue(r):
    for j in range (1, mcY + 1): 
        # reading cell value from source excel file 
        cY = wsY.cell(row = r, column = j)
        hY = wsY.cell(row = 1, column = j)
            
        c_valY = cY.value
        h_valY = hY.value
            
        if (h_valY == 'Colon_PTR'):
            y = c_valY

    return y

def getNumerator(r, c):
    cX = wsX.cell(row = r, column = c)
            
    x = cX.value
    y = getYvalue(i)

    num = x*y
    
    return num

def getDenumenator(r, c):
    # reading cell value from source excel file 
    cX = wsX.cell(row = r, column = c)
            
    x = cX.value

    dem = x*x
    return dem

# copying the cell values from source  
# excel file to destination excel file 
for j in range (1, mcX + 1): 
    for i in range (1, mrX + 1): 
        # reading cell value from source excel file 
        c = wsX.cell(row = i, column = j)
        h = wsX.cell(row = 1, column = j)

        c_val = c.value
        h_val = h.value

        if (j < 5):
            wsAI.cell(row = i, column = j).value = c.value 

        if (j > 4) and (i == 1):
            wsAI.cell(row = i, column = j).value = c.value 

        if (j > 4) and (i!=1):
            numerator = getNumerator(i,j)
            denominator = getDenumenator(i,j)

            if denominator == 0:
                ptrai = 0
            else:
                ptrai = numerator/denominator
            wsAI.cell(row = i, column = j).value = ptrai

wb.save(filename = 'Beta_Correlation.xlsx')
print("PTR_AI saved")

############################################################################################
## GENERATE BOXPLOT FROM DATA
############################################################################################
# import pandas
import pandas as pd
# import matplotlib
import matplotlib.pyplot as plt
# import seaborn
import seaborn as sns

df = pd.read_excel('Beta_Correlation.xlsx', sheet_name='PTR-AI')
print(df)

#df = df.drop(columns=['UAA', 'UGA','UAG'])
print(df)

# list(data)
header = list(df.columns) 

codons = []
trash = []
for i in range(0,64):
    if (i > 5) and (i < 64):
        codons.append(header[i])
    else: 
        trash.append(header[i]) 


data = pd.melt(df, id_vars=['GeneName'], value_vars= codons,
        var_name='Codons', value_name='value')
print(data)    

f = plt.figure(1)
sns.boxplot(y='value', x='Codons', 
                 data=data, 
                 palette="colorblind",
                 showfliers = False)


plt.rcParams["figure.figsize"] = (30, 8)
plt.ylabel(("PTR-log2(freq) Linear Regression Coefficient"))
plt.xticks(rotation=90)

plt.show()