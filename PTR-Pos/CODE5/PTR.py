
############################################################################################
# CALCULATE FREQUENCIES FOR CODONS IN SEQUENCES
############################################################################################

#import relevant libraries
import pandas as pd
import json 
import openpyxl as xl; 
  
# opening the source excel file 
wb1 = xl.load_workbook('Table_EV4.xlsx')
ws1 = wb1.worksheets[0] 
  
# opening the destination excel file  
# Create a Workbook
wb2 = xl.Workbook()
ws2 =  wb2.active
ws2.title = "Frequencies"
wb2.save(filename = 'PTRdata.xlsx')

wb2 = xl.load_workbook('PTRdata.xlsx') 
ws2 = wb2.active
  
# calculate total number of rows and  
# columns in source excel file 
mr = ws1.max_row 
mc = ws1.max_column 
  
# copying the cell values from source  
# excel file to destination excel file 
for i in range (1, mr + 1): 
    for j in range (1, mc + 1): 
        # reading cell value from source excel file 
        c = ws1.cell(row = i, column = j) 
  
        # writing the read value to destination excel file 
        ws2.cell(row = i, column = j).value = c.value 
  
# saving the destination excel file 
wb2.save(str('PTRdata.xlsx'))

#code for CodonTable fonction 
from Bio.Seq import Seq
from Bio.Alphabet import IUPAC
my_seq = Seq("GATCGATGGGCCTATATAGGATCGAAAATCGC", IUPAC.unambiguous_dna)

def CodonTable (my_seq):
    CodonsDict = { 
    "UUU": 0, "UUC": 0, "UUA": 0, "UUG": 0, "CUU": 0, 
    "CUC": 0, "CUA": 0, "CUG": 0, "AUU": 0, "AUC": 0, 
    "AUA": 0, "AUG": 0, "GUU": 0, "GUC": 0, "GUA": 0, 
    "GUG": 0, "UAU": 0, "UAC": 0, "UAA": 0, "UAG": 0, 
    "CAU": 0, "CAC": 0, "CAA": 0, "CAG": 0, "AAU": 0, 
    "AAC": 0, "AAA": 0, "AAG": 0, "GAU": 0, "GAC": 0, 
    "GAA": 0, "GAG": 0, "UCU": 0, "UCC": 0, "UCA": 0, 
    "UCG": 0, "CCU": 0, "CCC": 0, "CCA": 0, "CCG": 0, 
    "ACU": 0, "ACC": 0, "ACA": 0, "ACG": 0, "GCU": 0, 
    "GCC": 0, "GCA": 0, "GCG": 0, "UGU": 0, "UGC": 0, 
    "UGA": 0, "UGG": 0, "CGU": 0, "CGC": 0, "CGA": 0, 
    "CGG": 0, "AGU": 0, "AGC": 0, "AGA": 0, "AGG": 0, 
    "GGU": 0, "GGC": 0, "GGA": 0, "GGG": 0} 
    list_nucleotides =[my_seq[i:i+3] for i in range(0, len(my_seq), 3)]
    NumberCodons=0     
    for nucleotide in list_nucleotides:
        if not nucleotide in CodonsDict:
            continue

        else:
            CodonsDict[nucleotide]+= 1
            NumberCodons+=1
    return CodonsDict

#print(CodonTable(my_seq))

xls_file = pd.ExcelFile('PTRdata.xlsx') # Import the excel file and call it xls_file
df = xls_file.parse() #import into pandas dataframe object  

db_len = len(df)

gene_ids = df.get('CDS_Sequence')#only the CDS_Sequence column
output = pd.DataFrame()#empt7 df
for i in range(len(gene_ids)):
    seq = gene_ids[i]
    codon=CodonTable(seq)#use of CodontTable fonction that outputs a dictionary 
    output = output.append(codon, ignore_index=True)#append as rows with column the dictionary keys
#print(output.head())

df_merge_col = pd.merge(df, output, left_index=True, right_index=True)# use merge method, not join to add with respect to rows
df_merge_col.to_excel('PTRdata.xlsx',index=False)#use this method to save to csv, traditinal format, better than excel
'''
#############################################################################################################
#ADD PTR-AI SHEET TO EXCEL FILE
#############################################################################################################

from openpyxl import load_workbook
from decimal import Decimal
import math

wb = load_workbook('PTRdata.xlsx')

ws2 = wb.create_sheet("Sheet_PTRAI", 1)
ws2.title = "PTR-AI"

ws1 = wb.worksheets[0]
ws1.title = "Codon Frequencies" 

# calculate total number of rows and  
# columns in source excel file 
mr = ws1.max_row 
mc = ws1.max_column 

# copying the cell values from source  
# excel file to destination excel file 
for i in range (1, mr + 1): 
    for j in range (1, mc + 1): 
        # reading cell value from source excel file 
        c = ws1.cell(row = i, column = j)
        h = ws1.cell(row = 1, column = j)

        c_val = c.value
        h_val = h.value

        ws2.cell(row = i, column = j).value = c.value 

        if (j > 6) and (i!=1):
            freq = c_val
            if c_val == 0: #If there is no codon in sequence, ptr-ai = 0
                ptrai = 0
            else: # calculate PTR-AI values with equation: 10^log2(frequency)
                exp = math.log2(freq)
                r = math.pow(10, exp)
                result = math.log10(r)
                ptrai = round(result,2)
            ws2.cell(row = i, column = j).value = ptrai

wb.save(filename = 'PTRdata.xlsx')

###########################################################################################################
# CREATE BOXPLOT DATA FOR FIGURE
###########################################################################################################

import matplotlib.pyplot as plt
from openpyxl import load_workbook
import numpy as np
from matplotlib.patches import Polygon

wb = load_workbook('PTRdata.xlsx')
ws = wb.worksheets[1]

# calculate total number of rows and  
# columns in source excel file 
mr = ws.max_row 
mc = ws.max_column 

# copying the cell values from source  
# excel file to destination excel file 
dic = {}
for j in range (7, mc + 1): 
    ls = []
    h = ws.cell(row = 1, column = j)
    h_val = h.value

    name = h_val #get name of corresponding codon

    for i in range (2, mr + 1): 
        # reading cell value from source excel file 
        c = ws.cell(row = i, column = j)
        c_val = c.value
        
        ls.append(c_val) # input the data for each codon in a new list

    # Adding a new key value pair (dictionary of key = codon name; value = ptr-ai values list)
    dic.update( {name : ls} ) 

############################################################################################
## GENERATE BOXPLOT FROM DATA
############################################################################################

# split dictionary into keys and values 
keys = [] 
values = [] 
items = dic.items()

for item in items: 
    keys.append(item[0]), values.append(item[1])

# set ptr-ai values as values for the figure, and names of codons as the labels in x-axis
# remove outliers to make figure clearer
box_plot_data = values
plt.boxplot(box_plot_data, showfliers=False, patch_artist=True, labels = keys)
plt.rcParams["figure.figsize"] = (30, 8)

plt.ylabel(("PTR-AI log values"))
plt.title('PTR-AI Codon Values', fontsize=12)
plt.xticks(rotation=90)
plt.show()