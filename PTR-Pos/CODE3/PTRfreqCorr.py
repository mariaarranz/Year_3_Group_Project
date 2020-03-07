############################################################################################
# CALCULATE FREQUENCIES FOR CODONS IN SEQUENCES
############################################################################################

#import relevant libraries
import pandas as pd
import json 
import openpyxl as xl; 
  
# opening the source excel file 
wb1 = xl.load_workbook('Table_EV4.xlsx')
ws1 = wb1.worksheets[0] 
  
# opening the destination excel file  
# Create a Workbook
wb2 = xl.Workbook()
ws2 =  wb2.active
ws2.title = "Frequencies"
wb2.save(filename = 'PTR_FREQ_CODE3.xlsx')

wb2 = xl.load_workbook('PTR_FREQ_CODE3.xlsx') 
ws2 = wb2.active
  
# calculate total number of rows and  
# columns in source excel file 
mr = ws1.max_row 
mc = ws1.max_column 
  
# copying the cell values from source  
# excel file to destination excel file 
for i in range (1, mr + 1): 
    for j in range (1, mc + 1): 
        # reading cell value from source excel file 
        c = ws1.cell(row = i, column = j) 
  
        # writing the read value to destination excel file 
        ws2.cell(row = i, column = j).value = c.value 
  
# saving the destination excel file 
wb2.save(str('PTR_FREQ_CODE3.xlsx'))

#code for CodonTable fonction 
from Bio.Seq import Seq
from Bio.Alphabet import IUPAC
my_seq = Seq("GATCGATGGGCCTATATAGGATCGAAAATCGC", IUPAC.unambiguous_dna)

def CodonTable (my_seq):
    CodonsDict = { 
    "UUU": 0, "UUC": 0, "UUA": 0, "UUG": 0, "CUU": 0, 
    "CUC": 0, "CUA": 0, "CUG": 0, "AUU": 0, "AUC": 0, 
    "AUA": 0, "AUG": 0, "GUU": 0, "GUC": 0, "GUA": 0, 
    "GUG": 0, "UAU": 0, "UAC": 0, "UAA": 0, "UAG": 0, 
    "CAU": 0, "CAC": 0, "CAA": 0, "CAG": 0, "AAU": 0, 
    "AAC": 0, "AAA": 0, "AAG": 0, "GAU": 0, "GAC": 0, 
    "GAA": 0, "GAG": 0, "UCU": 0, "UCC": 0, "UCA": 0, 
    "UCG": 0, "CCU": 0, "CCC": 0, "CCA": 0, "CCG": 0, 
    "ACU": 0, "ACC": 0, "ACA": 0, "ACG": 0, "GCU": 0, 
    "GCC": 0, "GCA": 0, "GCG": 0, "UGU": 0, "UGC": 0, 
    "UGA": 0, "UGG": 0, "CGU": 0, "CGC": 0, "CGA": 0, 
    "CGG": 0, "AGU": 0, "AGC": 0, "AGA": 0, "AGG": 0, 
    "GGU": 0, "GGC": 0, "GGA": 0, "GGG": 0} 
    list_nucleotides =[my_seq[i:i+3] for i in range(0, len(my_seq), 3)]
    NumberCodons=0     
    for nucleotide in list_nucleotides:
        if not nucleotide in CodonsDict:
            continue

        else:
            CodonsDict[nucleotide]+= 1
            NumberCodons+=1
    return CodonsDict

#print(CodonTable(my_seq))

xls_file = pd.ExcelFile('PTR_FREQ_CODE3.xlsx') # Import the excel file and call it xls_file
df = xls_file.parse() #import into pandas dataframe object  

db_len = len(df)

gene_ids = df.get('CDS_Sequence')#only the CDS_Sequence column
output = pd.DataFrame()#empt7 df
for i in range(len(gene_ids)):
    seq = gene_ids[i]
    codon=CodonTable(seq)#use of CodontTable fonction that outputs a dictionary 
    output = output.append(codon, ignore_index=True)#append as rows with column the dictionary keys
#print(output.head())

df_merge_col = pd.merge(df, output, left_index=True, right_index=True)# use merge method, not join to add with respect to rows
df_merge_col.to_excel('PTR_FREQ_CODE3.xlsx',index=False)#use this method to save to csv, traditinal format, better than excel

###########################################################################################################
## MAKE NEW SHEET WITH PTR values FOR EACH GENE
###########################################################################################################

import pandas as pd
import json 
import openpyxl as xl 
from openpyxl import load_workbook
from decimal import Decimal
import math

wb = load_workbook('PTR_FREQ_CODE3.xlsx')

wb = load_workbook('PTR_FREQ_CODE3.xlsx')

ws1 = wb.worksheets[0]
ws1.title = "Codon Frequencies" 

ws3 = wb.create_sheet("Sheet_Y", 1)
ws3.title = "PTR"
  
# opening the source excel file 
wbEV3 = xl.load_workbook('Table_EV3.xlsx')
wsEV3 = wbEV3.worksheets[0] 

# calculate total number of rows and  
# columns in source excel file 
mr = wsEV3.max_row 
mc = wsEV3.max_column 

def getYbar(col):
    ls = []
    for i in range (2, 11577): 
        # reading cell value from source excel file 
        c = wsEV3.cell(row = i, column = col)
        c_val = c.value
    
        ls.append(c_val)
    
    summ = 0
    for n in ls:
        summ  += n

    ybar = float(summ/11574)
    return ybar
    
# copying the cell values from source  
# excel file to destination excel file 
for j in range (1, mc + 1): 
    for i in range (1, mr + 1): 
        # reading cell value from source excel file 
        c = wsEV3.cell(row = i, column = j)
        h = wsEV3.cell(row = 1, column = j)

        c_val = c.value
        h_val = h.value

        if (j < 5):
            ws3.cell(row = i, column = j).value = c.value 

        if (j > 4) and (i == 1):
            ws3.cell(row = i, column = j).value = c.value 

        if (j > 4) and (i!=1):
            y = float(c_val)
            ws3.cell(row = i, column = j).value = y

wb.save(filename = 'PTR_FREQ_CODE3.xlsx')


###########################################################################################################
# CREATE BOXPLOT DATA FOR FIGURE
###########################################################################################################

# import pandas
import pandas as pd
# import matplotlib
import matplotlib.pyplot as plt
# import seaborn
import seaborn as sns

df = pd.read_excel('PTR_FREQ_CODE3.xlsx', sheet_name='Codon Frequencies')
#print(df)

dfY = pd.read_excel('PTR_FREQ_CODE3.xlsx', sheet_name='PTR')
#print(dfY)

dfPTR = dfY['Liver_PTR']
#print(dfPTR)

df['PTR'] = dfPTR
df = df.drop(columns=['EnsemblGeneID', 'EnsemblTranscriptID', 'EnsemblProteinID', 'Unnamed: 4'])
print(df)


#Reference - https://levelup.gitconnected.com/pearson-coefficient-of-correlation-using-pandas-ca68ce678c04

corr = df.corr(method='pearson')
print(corr)

f = plt.figure(1)
sns.heatmap(corr, 
            xticklabels=corr.columns,
            yticklabels=corr.columns,
            cmap='RdBu_r',
            annot=False,
            linewidth=0.5)


corr = corr.drop(['PTR'])
print(corr)

# list(data)
header = list(corr.columns) 

codons = []
trash = []
for i in range(0,64):
    codons.append(header[i])

corr = corr.drop(columns=codons). assign(Codons = codons)
print(corr)

g = plt.figure(2)
sns.scatterplot(x="Codons", y="PTR", data=corr)

xs = corr.Codons
ys = corr.PTR

for x,y in zip(xs,ys):

    label1 = "{}".format(x)
    label2 = ", {:.2f}".format(y)
    label = label1 + label2

    plt.annotate(label, # this is the text
                 (x,y), # this is the point to label
                 textcoords="offset points", # how to position the text
                 xytext=(0,10), # distance from text to points (x,y)
                 ha='center') # horizontal alignment can be left, right or center


plt.rcParams["figure.figsize"] = (30, 8)
plt.ylabel(("Pearson Correlation Coefficient"))
plt.title('PTR to Codon Frequency Pearson Correlation', fontsize=12)
plt.xticks(rotation=90)

plt.show()
