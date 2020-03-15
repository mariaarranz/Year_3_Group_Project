########################################################################################################
# Create Workbook and add Codon Frequencies Table
###########################################################################################################

from decimal import Decimal
import math
from openpyxl import Workbook

wb = Workbook()
wb.save(filename='CodonPTR.xlsx')

ws2 = wb.create_sheet("Codons", 0)
ws2.title = "Codon Frequencies"

from openpyxl import load_workbook
wb1 = load_workbook('CodonTable.xlsx')
ws1 = wb1.worksheets[0]

# calculate total number of rows and  
# columns in source excel file 
mr = ws1.max_row 
mc = ws1.max_column 

# copying the cell values from source  
# excel file to destination excel file 
for i in range (1, mr + 1): 
    for j in range (1, mc + 1): 
        # reading cell value from source excel file 
        c = ws1.cell(row = i, column = j)
        h = ws1.cell(row = 1, column = j)

        c_val = c.value
        h_val = h.value

        ws2.cell(row = i, column = j).value = c.value 

wb.save(filename = 'CodonPTR.xlsx')
###########################################################################################################
## Make column of PTR for SPECIFIC TISSUE (BRAIN)
###########################################################################################################

import pandas as pd
import json 
import openpyxl as xl 
from openpyxl import load_workbook
from decimal import Decimal
import math

wb = load_workbook('CodonPTR.xlsx')

ws1 = wb.worksheets[0]
ws1.title = "Codon Frequencies" 
  
# opening the source excel file 
wbEV3 = xl.load_workbook('Table_EV3.xlsx')
wsEV3 = wbEV3.worksheets[0] 

# calculate total number of rows and  
# columns in source excel file 
mrEV3 = wsEV3.max_row 
mcEV3 = wsEV3.max_column 

mr = ws1.max_row 
mc = ws1.max_column
    
# copying the cell values from source  
# excel file to destination excel file 
for j in range (1, mcEV3 + 1):
    for i in range (1, mrEV3 + 1): 
        # reading cell value from source excel file 
        c = wsEV3.cell(row = i, column = j)
        h = wsEV3.cell(row = 1, column = j)

        c_val = c.value
        h_val = h.value

        if (h_val == 'Brain_PTR'):
            ws1.cell(row = i, column = mc+1).value = c.value 

wb.save(filename = 'CodonPTR.xlsx')

############################################################################################
## GENERATE BOXPLOT FROM DATA
############################################################################################

# import pandas
import pandas as pd
# import matplotlib
import matplotlib.pyplot as plt
# import seaborn
import seaborn as sns

df = pd.read_excel('CodonPTR.xlsx', sheet_name='Codon Frequencies')
print(df)

dfTOP = df.sort_values(by ='Brain_PTR', ascending=False).head(int(df.shape[0]*.05)).assign(Location='Top 5%')
#print(dfTOP)

dfBOT = df.sort_values(by ='Brain_PTR', ascending=True).head(int(df.shape[0]*.05)).assign(Location='Bottom 5%')
#print(dfBOT)

# list(data)
dfheader = df.drop(columns=['EnsemblGeneID', 'EnsemblTranscriptID', 'EnsemblProteinID', 'GeneName', 'Brain_PTR'])
codons = list(dfheader.columns) 
print(codons)

cdf = pd.concat([dfTOP, dfBOT])
print(cdf)    

data = pd.melt(cdf, id_vars=['Location'], value_vars= codons)
print(data)    

sns.boxplot(y='value', x='variable', 
                 data=data, 
                 palette="colorblind",
                 hue='Location',
                 showfliers = False)


plt.rcParams["figure.figsize"] = (30, 8)
plt.ylabel(("Codon Frequencies"))
plt.xlabel(("Codons"))
plt.xticks(rotation=90)
plt.show()