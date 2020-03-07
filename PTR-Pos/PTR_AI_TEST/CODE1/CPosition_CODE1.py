# CODE FOR POSITION CODON vs PTR-AI VALUE

############################################################################################
# CALCULATE THE POSITION OF THE SPECIFIED CODON IN THE SEQUENCES
############################################################################################

#import relevant libraries
import pandas as pd
import openpyxl as xl; 

# opening the source excel file 
wb1 = xl.load_workbook('Table_EV4.xlsx')
ws1 = wb1.worksheets[0] 
  
# opening the destination excel file  
# Create a Workbook
wb2 = xl.Workbook()
ws2 =  wb2.active
ws2.title = "Positions"
wb2.save(filename = 'PosCODE1.xlsx')

wb2 = xl.load_workbook('PosCODE1.xlsx') 
ws2 = wb2.active
  
# calculate total number of rows and  
# columns in source excel file 
mr = ws1.max_row 
mc = ws1.max_column 
  
# copying the cell values from source  
# excel file to destination excel file 
for i in range (1, mr + 1): 
    for j in range (1, mc + 1): 
        # reading cell value from source excel file 
        c = ws1.cell(row = i, column = j) 
  
        # writing the read value to destination excel file 
        ws2.cell(row = i, column = j).value = c.value 
  
# saving the destination excel file 
wb2.save(str('PosCODE1.xlsx'))

#code for CodonPosition function 
from Bio.Seq import Seq
from Bio.Alphabet import IUPAC
my_seq = "GAUCGAUGGGCUUAUAUAGGAUCGAAAAUCGCA"

aa = 'AAA' # select codon to find data for
def CodonPosition (my_seq):
    codons = []
    start = 0
    # separate the sequences into a codons list
    for l in range(int(len(my_seq)/3)):
        end = start + 3
        codon = my_seq[start:end]
        codons.append(codon)
        start = end
    #print(codons)
    
    # Find only the position of the selected codon
    if aa not in codons:
        pos = '00' #if not in sequence, output '00'
    else:
        pos = codons.index(aa) #if in sequence, output the position on the list

    PosDict = {aa: pos} #create a dictionary for the codon and its positions
    
    return PosDict

#print(CodonPosition(my_seq))

xls_file = pd.ExcelFile('PosCODE1.xlsx') # Import the excel file and call it xls_file
df = xls_file.parse() #import into pandas dataframe object  

db_len = len(df)

gene_ids = df.get('CDS_Sequence')#only the CDS_Sequence column
output = pd.DataFrame()#empt7 df
for i in range(len(gene_ids)):
    seq = gene_ids[i]
    codon = CodonPosition(seq)#use of CodontTable fonction that outputs a dictionary 
    output = output.append(codon, ignore_index=True)#append as rows with column the dictionary keys
#print(output.head())

df_merge_col = pd.merge(df, output, left_index=True, right_index=True)# use merge method, not join to add with respect to rows
df_merge_col.to_excel('PosCODE1.xlsx',index=False)#use this method to save to csv, traditinal format, better than excel

####################################################################################
# GET PTR-AI DATA and GENERATE FIGURE
####################################################################################
#ADD PTR-AI SHEET TO EXCEL FILE
from openpyxl import load_workbook
from decimal import Decimal
import math

wb = load_workbook('PosCODE1.xlsx')

ws2 = wb.create_sheet("Sheet_PTRAI", 1)
ws2.title = "PTR-AI"

ws1 = wb.worksheets[0]
ws1.title = "Positions" #rename sheet 0

wbP = load_workbook('PTRCODE1.xlsx')
wsP = wbP.worksheets[3]

# calculate total number of rows and  
# columns in source excel file 
mr = wsP.max_row 
mc = wsP.max_column 

# copying the cell values from source  
# excel file to destination excel file 
for i in range (1, mr + 1): 
    for j in range (1, mc + 1): 
        # reading cell value from Pos.xlsx file 
        # get postion data
        c = wsP.cell(row = i, column = j)
        h = wsP.cell(row = 1, column = j)

        c_val = c.value
        h_val = h.value

        if (j < 5): 
            ws2.cell(row = i, column = j).value = c.value

        if (j == 4) and (i == 1):
            ws2.cell(row = i, column = j).value = aa

        # Create new sheet in Pos.xlsx file containing the PTR-AI values for 
        # the selected codon
        if (i != 1):
            aa_col = 'Codon_'+ aa
            if (h_val == aa_col):
                ws2.cell(row = i, column = 4).value = c.value

wb.save(filename = 'PosCODE1.xlsx')

#########################################################################################
# CREATE DATA FOR FIGURE
#########################################################################################

import matplotlib.pyplot as plt
from openpyxl import load_workbook
import numpy as np
from matplotlib.patches import Polygon

wb = load_workbook('PosCODE1.xlsx')
ws0 = wb.worksheets[0] # sheet containing postions
ws1 = wb.worksheets[1] # sheet containing ptr-ai values

# calculate total number of rows and  
# columns in source excel file 
mr = ws0.max_row 
mc = ws0.max_column 

mr1 = ws1.max_row 
mc1 = ws1.max_column 

# copying the cell values from source  
# excel file to destination excel file 
aa = 'AAA'
pos_ls = []
ptrai_ls = []

ls1 = [] #create list of all positions
for i in range (2, mr + 1): 
    # reading cell value from source excel file 
    c = ws0.cell(row = i, column = 7)
    c_val = c.value
    ls1.append(c_val)
# Position values list
pos_ls = ls1

ls2 = []
for i in range (2, mr1 + 1): 
    # reading cell value from source excel file 
    c = ws1.cell(row = i, column = 4)
    c_val = c.value
    ls2.append(c_val)
# PTR-AI values list
ptrai_ls = ls2

#########################################################################################
# GENERATE SCATTER_PLOT
#########################################################################################
import matplotlib.pyplot as plt 

# x-axis values 
x = pos_ls 
# y-axis values 
y = ptrai_ls
  
# plotting points as a scatter plot 
plt.scatter(x, y, label= "stars", color= "green",  
            marker= "*", s=10) 
  
# x-axis label 
plt.xlabel('Codon ' + aa + ' postions') 
# frequency label 
plt.ylabel('PTR-AI values') 
# plot title 
plt.title('Position Relevance of Codon '+ aa) 
# showing legend 
plt.legend() 
  
# function to show the plot 
plt.show() 
