# CODE FOR POSITION CODON vs PTR-AI VALUE

############################################################################################
# CALCULATE THE POSITION OF THE SPECIFIED CODON IN THE SEQUENCES

#import relevant libraries
import pandas as pd
import requests, sys #to communicate with Ensembl servers
import json 
import openpyxl as xl; 

# opening the source excel file 
wb1 = xl.load_workbook('Table_EV4.xlsx')
ws1 = wb1.worksheets[0] 
  
# opening the destination excel file  
# Create a Workbook
wb2 = xl.Workbook()
ws2 =  wb2.active
ws2.title = "Positions"
wb2.save(filename = 'Pos.xlsx')

wb2 = xl.load_workbook('Pos.xlsx') 
ws2 = wb2.active
  
# calculate total number of rows and  
# columns in source excel file 
mr = ws1.max_row 
mc = ws1.max_column 
  
# copying the cell values from source  
# excel file to destination excel file 
for i in range (1, mr + 1): 
    for j in range (1, mc + 1): 
        # reading cell value from source excel file 
        c = ws1.cell(row = i, column = j) 
  
        # writing the read value to destination excel file 
        ws2.cell(row = i, column = j).value = c.value 
  
# saving the destination excel file 
wb2.save(str('Pos.xlsx'))

#code for CodonTable fonction 
from Bio.Seq import Seq
from Bio.Alphabet import IUPAC
my_seq = "GAUCGAUGGGCUUAUAUAGGAUCGAAAAUCGCA"

aa = 'AAA'
def CodonPosition (my_seq):
    codons = []
    start = 0
    for l in range(int(len(my_seq)/3)):
        end = start + 3
        codon = my_seq[start:end]
        codons.append(codon)
        start = end
    #print(codons)
    
    if aa not in codons:
        pos = '00'
    else:
        pos = codons.index(aa)

    PosDict = {aa: pos}
    
    return PosDict

#print(CodonPosition(my_seq))

xls_file = pd.ExcelFile('Pos.xlsx') # Import the excel file and call it xls_file
df = xls_file.parse() #import into pandas dataframe object  

db_len = len(df)

gene_ids = df.get('CDS_Sequence')#only the CDS column
output = pd.DataFrame()#empt7 df
for i in range(len(gene_ids)):
    seq = gene_ids[i]
    codon = CodonPosition(seq)#use of CodontTable fonction that outputs a dictionary 
    output = output.append(codon, ignore_index=True)#append as rows with column the dictionary keys
#print(output.head())

df_merge_col = pd.merge(df, output, left_index=True, right_index=True)# use merge method, not join to add with respect to rows
df_merge_col.to_excel('Pos.xlsx',index=False)#use this method to save to csv, traditinal format, better than excel

####################################################################################
# GET PTR-AI DATA and GENERATE FIGURE
#############################################################################################################
#ADD PTR-AI SHEET TO EXCEL FILE
from openpyxl import load_workbook
from decimal import Decimal
import math

wb = load_workbook('Pos.xlsx')

ws2 = wb.create_sheet("Sheet_PTRAI", 1)
ws2.title = "PTR-AI"

ws1 = wb.worksheets[0]
ws1.title = "Positions" 

wbP = load_workbook('PTRdata.xlsx')
wsP = wbP.worksheets[1]

# calculate total number of rows and  
# columns in source excel file 
mr = ws1.max_row 
mc = ws1.max_column 

# copying the cell values from source  
# excel file to destination excel file 
for i in range (1, mr + 1): 
    for j in range (1, mc + 1): 
        # reading cell value from source excel file 
        c = ws1.cell(row = i, column = j)
        h = ws1.cell(row = 1, column = j)

        c_val = c.value
        h_val = h.value

        ws2.cell(row = i, column = j).value = c.value

        cp = wsP.cell(row = i, column = j)
        hp = wsP.cell(row = 1, column = j)
        
        cP_val = cp.value
        hP_val = hp.value

        if (j == 7) and (i != 1):
            if (hP_val == aa):
                ws2.cell(row = i, column = j).value = cp.value

wb.save(filename = 'Pos.xlsx')

#########################################################################################
# CREATE DATA FOR FIGURE
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import numpy as np
from matplotlib.patches import Polygon

wb = load_workbook('Pos.xlsx')
ws0 = wb.worksheets[0]
ws1 = wb.worksheets[1]

# calculate total number of rows and  
# columns in source excel file 
mr = ws0.max_row 
mc = ws0.max_column 

# copying the cell values from source  
# excel file to destination excel file 
aa = 'AAA'
pos_ls = []
ptrai_ls = []
for j in range (7, mc + 1): 
    ls1 = []
    for i in range (2, mr + 1): 
        # reading cell value from source excel file 
        c = ws0.cell(row = i, column = j)
        c_val = c.value
        ls1.append(c_val)
    # Adding a new key value pair
    pos_ls = ls1

    ls2 = []
    for i in range (2, mr + 1): 
        # reading cell value from source excel file 
        c = ws1.cell(row = i, column = j)
        c_val = c.value
        ls2.append(c_val)
    # Adding a new key value pair
    ptrai_ls = ls2

#print(pos_ls)
#print(ptrai_ls)

##########################################################################################
# GENERATE SCATTER_PLOT
import matplotlib.pyplot as plt 

# x-axis values 
x = pos_ls 
# y-axis values 
y = ptrai_ls
  
# plotting points as a scatter plot 
plt.scatter(x, y, label= "stars", color= "green",  
            marker= "*", s=10) 
  
# x-axis label 
plt.xlabel('Codon ' + aa + ' postions') 
# frequency label 
plt.ylabel('PTR-AI values') 
# plot title 
plt.title('Position Relevance of Codon '+ aa) 
# showing legend 
plt.legend() 
  
# function to show the plot 
plt.show() 
