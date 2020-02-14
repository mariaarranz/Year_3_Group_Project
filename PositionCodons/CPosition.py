# CODE FOR POSITION CODON vs PTR-AI VALUE

#import relevant libraries
import pandas as pd
import requests, sys #to communicate with Ensembl servers
import json 
import openpyxl as xl; 
  
# opening the source excel file 
wb1 = xl.load_workbook('Table_EV4.xlsx')
ws1 = wb1.worksheets[0] 
  
# opening the destination excel file  
# Create a Workbook
wb2 = xl.Workbook()
ws2 =  wb2.active
ws2.title = "Positions"
wb2.save(filename = 'Pos.xlsx')

wb2 = xl.load_workbook('Pos.xlsx') 
ws2 = wb2.active
  
# calculate total number of rows and  
# columns in source excel file 
mr = ws1.max_row 
mc = ws1.max_column 
  
# copying the cell values from source  
# excel file to destination excel file 
for i in range (1, mr + 1): 
    for j in range (1, mc + 1): 
        # reading cell value from source excel file 
        c = ws1.cell(row = i, column = j) 
  
        # writing the read value to destination excel file 
        ws2.cell(row = i, column = j).value = c.value 
  
# saving the destination excel file 
wb2.save(str('Pos.xlsx'))

#code for CodonTable fonction 
from Bio.Seq import Seq
from Bio.Alphabet import IUPAC
my_seq = "GAUCGAUGGGCUUAUAUAGGAUCGAAAAUCGCA"

def CodonPosition (my_seq):
    codons = []
    pos_ls = []
    start = 0
    for l in range(int(len(my_seq)/3)):
        end = start + 3
        codon = my_seq[start:end]
        codons.append(codon)
        start = end
    #print(codons)
    
    c = 'AAA'
    if c not in codons:
        pos = '---'
    else:
        pos = codons.index(c)

    PosDict = {c: pos}
    
    return PosDict

#print(CodonPosition(my_seq))

xls_file = pd.ExcelFile('Pos.xlsx') # Import the excel file and call it xls_file
df = xls_file.parse() #import into pandas dataframe object  

db_len = len(df)

gene_ids = df.get('CDS_Sequence')#only the CDS column
output = pd.DataFrame()#empt7 df
for i in range(len(gene_ids)):
    seq = gene_ids[i]
    codon = CodonPosition(seq)#use of CodontTable fonction that outputs a dictionary 
    output = output.append(codon, ignore_index=True)#append as rows with column the dictionary keys
#print(output.head())

df_merge_col = pd.merge(df, output, left_index=True, right_index=True)# use merge method, not join to add with respect to rows
df_merge_col.to_excel('Pos.xlsx',index=False)#use this method to save to csv, traditinal format, better than excel
