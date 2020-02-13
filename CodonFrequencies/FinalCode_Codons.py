# -*- coding: utf-8 -*-
"""
Year 3 Project: Codons
Code to extract protein ID from table and 
retrieve the cds's from the Ensembl server.
Careful: loops through whole table so only use once on college computer!
Created on Mon Dec  9 17:54:22 2019
@author: Marianne Buat
"""
#import relevant libraries

import pandas as pd
import requests, sys #to communicate with Ensembl servers
import json 

xls_file = pd.ExcelFile('Table_EV1.xlsx') # Import the excel file and call it xls_file
df = xls_file.parse('C. Genes') #import into pandas only one of the sheets  

db_len = len(df)
#Create an output dataframe with only the parts we are interested in
geneIDs = df.get('Gene ID')
CDS=pd.Series([]) #Create new series to contain the sequences, will then be added to output dataframe
actualID=pd.Series([])

maxQueryLen = 10 #actually 50
server = "https://rest.ensembl.org"
ext = "/sequence/id"
headers={ "Content-Type" : "application/json", "Accept" : "application/json","type":'cds'}


#Loop through huge table to retrieve all sequences
ind = 0
#while ind+maxQueryLen<=db_len:
gene_ids = geneIDs[ind:ind+maxQueryLen+1] #select 50 IDs from table
id_list=gene_ids.values.tolist() #convert dataframe slice to list
 #convert to json string for compatibility (could also have used str())

inputData = json.dumps({'ids':id_list,'type':['cds']*len(id_list)}) #dictionary with ids and output type cds for all; 

r = requests.post(server+ext, headers=headers, data=inputData) #request data from server        
if not r.ok:
  r.raise_for_status()
  sys.exit()
 
decoded = r.json() #read response from server
#decoded is a list of dictionaries
for k in range(len(id_list)):
 CDS[ind+k] = decoded[k]['seq'] #store the retrieved sequences
 actualID[ind+k] = decoded[k]['id']

#Create an output dataframe for the results

outputDF = pd.DataFrame({'Gene ID':geneIDs,'answer ID':actualID,'CDS':CDS})
print(outputDF.head())                                                                

# #Save to file
# writer = pd.ExcelWriter('CDSQueryOutput.xlsx')
# outputDF.to_excel(writer)
# writer.save()
# print('Output is written successfully to Excel File.')

##
#code for CodonTable fonction 
from Bio.Seq import Seq
from Bio.Alphabet import IUPAC
my_seq = Seq("GATCGATGGGCCTATATAGGATCGAAAATCGC", IUPAC.unambiguous_dna)

def CodonTable (my_seq):
    CodonsDict = { 
    "TTT": 0, "TTC": 0, "TTA": 0, "TTG": 0, "CTT": 0, 
    "CTC": 0, "CTA": 0, "CTG": 0, "ATT": 0, "ATC": 0, 
    "ATA": 0, "ATG": 0, "GTT": 0, "GTC": 0, "GTA": 0, 
    "GTG": 0, "TAT": 0, "TAC": 0, "TAA": 0, "TAG": 0, 
    "CAT": 0, "CAC": 0, "CAA": 0, "CAG": 0, "AAT": 0, 
    "AAC": 0, "AAA": 0, "AAG": 0, "GAT": 0, "GAC": 0, 
    "GAA": 0, "GAG": 0, "TCT": 0, "TCC": 0, "TCA": 0, 
    "TCG": 0, "CCT": 0, "CCC": 0, "CCA": 0, "CCG": 0, 
    "ACT": 0, "ACC": 0, "ACA": 0, "ACG": 0, "GCT": 0, 
    "GCC": 0, "GCA": 0, "GCG": 0, "TGT": 0, "TGC": 0, 
    "TGA": 0, "TGG": 0, "CGT": 0, "CGC": 0, "CGA": 0, 
    "CGG": 0, "AGT": 0, "AGC": 0, "AGA": 0, "AGG": 0, 
    "GGT": 0, "GGC": 0, "GGA": 0, "GGG": 0} 
    list_nucleotides =[my_seq[i:i+3] for i in range(0, len(my_seq), 3)]
    NumberCodons=0     
    for nucleotide in list_nucleotides:
        if not nucleotide in CodonsDict:
            continue

        else:
            CodonsDict[nucleotide]+= 1
            NumberCodons+=1
    return CodonsDict

#print(CodonTable(my_seq))

import pandas as pd
import requests, sys #to communicate with Ensembl servers
import json 


xls_file = pd.ExcelFile('CDSOutput.xlsx') # Import the excel file and call it xls_file
df = xls_file.parse() #import into pandas dataframe object  

db_len = len(df)

gene_ids = df.get('CDS')#only the CDS column
output = pd.DataFrame()#empt7 df
for i in range(len(gene_ids)):
    seq = gene_ids[i]
    codon=CodonTable(seq)#use of CodontTable fonction that outputs a dictionary 
    output = output.append(codon, ignore_index=True)#append as rows with column the dictionary keys
    k=k+1
#print(output.head())

df_merge_col = pd.merge(df, output, left_index=True, right_index=True)# use merge method, not join to add with respect to rows
df_merge_col.to_excel('CodonTable.xlsx',index=False)#use this method to save to csv, traditinal format, better than excel


#ADD RSCU SHEET TO EXCEL FILE
from openpyxl import load_workbook
from decimal import Decimal

wb = load_workbook('CodonTable.xlsx')

ws2 = wb.create_sheet("Sheet_RSCU", 1)
ws2.title = "RSCU"

ws1 = wb.worksheets[0]
ws1.title = "Codon Frequencies" 

# calculate total number of rows and  
# columns in source excel file 
mr = ws1.max_row 
mc = ws1.max_column 

#RELATIVE SYNONYMOUS CODON USAGE CALCULATION
translation_list = {'*':['TAA','TAG','TGA'],
                    'A':['GCA','GCC','GCG','GCT'],
                    'C':['TGC','TGT'],
                    'D':['GAC','GAT'],
                    'E':['GAA','GAG'],
                    'F':['TTC','TTT'],
                    'G':['GGA','GGC','GGG','GGT'],
                    'H':['CAC','CAT'],
                    'I':['ATA','ATC','ATT'],
                    'K':['AAA','AAG'],
                    'L':['CTA','CTC','CTG','CTT','TTA','TTG'],
                    'M':['ATG'],
                    'N':['AAC','AAT'],
                    'P':['CCA','CCC','CCG','CCT'],
                    'Q':['CAA','CAG'],
                    'R':['AGA','AGG','CGA','CGC','CGG','CGT'],
                    'S':['AGC','AGT','TCA','TCC','TCG','TCT'],
                    'T':['ACA','ACC','ACG','ACT'],
                    'V':['GTA','GTC','GTG','GTT'],
                    'W':['TGG'],
                    'Y':['TAC','TAT']}


# copying the cell values from source  
# excel file to destination excel file 
for i in range (1, mr + 1): 
    codon_dict = {}
    rscu = {}
    for j in range (1, mc + 1): 
        # reading cell value from source excel file 
        c = ws1.cell(row = i, column = j)
        h = ws1.cell(row = 1, column = j)

        c_val = c.value
        h_val = h.value
        
        ws2.cell(row = i, column = j).value = c.value 
        
        #RSCU values are CodonCount/((1/num of synonymous codons) * sum of all synonymous codons)
        if (i>1) and (j>5):
            for aa in translation_list:
                s_codons = translation_list[aa]
                
                for s in s_codons:
                    count = c_val
                    codon_dict[h_val] = count

    #print(codon_dict)
    
    for aa in translation_list:
        total = 0.0
        s_codons = translation_list[aa]

        #THIS WILL COUNT HOW MANY TIMES EACH CODON APPEARS IN THE SEQUENCE
        for s in s_codons:
            count = codon_dict.get(s)
            if type(count) == int:
                total += count

        #calculate the RSCU value for each of the codons
        for  s in s_codons:
            count = codon_dict.get(s)
            if type(count) == int:
                denominator = float(total)/ len(s_codons)
                if denominator == 0:
                    r = 0.0
                else:
                    r = Decimal(count/denominator)
                result = round(r,2)
                rscu[s] = result
        
    final_rscu = {}
    for key in codon_dict:
        value = rscu.get(key)
        final_rscu[key] = value
    
    #print('\n')
    #print(final_rscu)

    for j in range (5, mc + 1):
        if (i != 1):
            # reading cell value from source excel file 
            c = ws1.cell(row = i, column = j)
            h = ws1.cell(row = 1, column = j)

            c_val = c.value
            h_val = h.value
        
            ws2.cell(row = i, column = j).value = final_rscu.get(h_val) 

wb.save(filename = 'CodonTable.xlsx')
####
## Calculate CAI value from CAI package
from CAI import CAI
from Bio import SeqIO # to parse FASTA files

ws3 = wb.create_sheet("Sheet_CAI", 2)
ws3.title = "CAI"

ws3.cell(row = 1, column = 5).value = "CAI"

# calculate total number of rows and  
# columns in source excel file 
mr = ws1.max_row 
mc = ws1.max_column 

# copying the cell values from source  
# excel file to destination excel file 
for i in range (1, mr + 1): 
    codon_dict = {}
    rscu = {}
    for j in range (1, 5): 
        # reading cell value from source excel file 
        c = ws1.cell(row = i, column = j)
        h = ws1.cell(row = 1, column = j)

        c_val = c.value
        h_val = h.value
        
        ws3.cell(row = i, column = j).value = c.value 
        
        if (j == 4) and (i!=1):
            sequence = c_val
            reference = [seq.seq for seq in SeqIO.parse("reference.fasta", "fasta")]
            if 'N' in sequence:
                cai = ' '

            elif len(sequence) % 3 != 0:
                cai = ' '
            else:
                cai = CAI(sequence, reference=reference)
            ws3.cell(row = i, column = j+1).value = cai

wb.save(filename = 'CodonTable.xlsx')