from CAI import CAI
from Bio import SeqIO # to parse FASTA files

from openpyxl import load_workbook

wb = load_workbook('CodonTable.xlsx')

ws3 = wb.create_sheet("Sheet_RSCU", 2)
ws3.title = "CAI"

ws3.cell(row = 1, column = 5).value = "CAI"

ws1 = wb.worksheets[0]

# calculate total number of rows and  
# columns in source excel file 
mr = ws1.max_row 
mc = ws1.max_column 

# copying the cell values from source  
# excel file to destination excel file 
for i in range (1, mr + 1): 
    codon_dict = {}
    rscu = {}
    for j in range (1, 5): 
        # reading cell value from source excel file 
        c = ws1.cell(row = i, column = j)
        h = ws1.cell(row = 1, column = j)

        c_val = c.value
        h_val = h.value
        
        ws3.cell(row = i, column = j).value = c.value 
        
        if (j == 4) and (i!=1):
            sequence = c_val
            reference = [seq.seq for seq in SeqIO.parse("reference.fasta", "fasta")]
            if 'N' in sequence:
                cai = ' '

            """
            For some reason some of the sequences in sequence contain
            a letter 'N'

            Get Error -->   Traceback (most recent call last):
                            File "C:\Users\maria\AppData\Local\Programs\Python\Python37\lib\site-packages\CAI\CAI.py", line 204, in CAI
                                sequence_weights.append(weights[codon])
                            KeyError: 'NNG'

                            During handling of the above exception, another exception occurred:

                            Traceback (most recent call last):
                            File "CAImini.py", line 38, in <module>
                                cai = CAI(sequence, reference=reference)
                            File "C:\Users\maria\AppData\Local\Programs\Python\Python37\lib\site-packages\CAI\CAI.py", line 210, in CAI
                                raise KeyError("Bad weights dictionary passed: missing weight for codon.")
                            KeyError: 'Bad weights dictionary passed: missing weight for codon.'
            """

            elif len(sequence) % 3 != 0:
                cai = ' '

            """
            For some reason some of the sequences in sequence cannot
            be divisible by 3 (cannot be divided into its codons correctly)

            Get Error -->   Traceback (most recent call last):
                            File "CAImini.py", line 41, in <module>
                                cai = CAI(sequence, reference=reference)
                            File "C:\Users\maria\AppData\Local\Programs\Python\Python37\lib\site-packages\CAI\CAI.py", line 186, in CAI
                                raise ValueError("Input sequence not divisible by three")
                            ValueError: Input sequence not divisible by three
            """
            else:
                cai = CAI(sequence, reference=reference)
            print(cai)
            ws3.cell(row = i, column = j+1).value = cai

wb.save(filename = 'CodonTable.xlsx')

