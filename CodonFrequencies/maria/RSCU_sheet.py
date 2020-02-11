from openpyxl import load_workbook
from decimal import Decimal

wb = load_workbook('CodonTable.xlsx')

ws2 = wb.create_sheet("Sheet_RSCU", 1)
ws2.title = "RSCU"

ws1 = wb.worksheets[0] 

# calculate total number of rows and  
# columns in source excel file 
mr = ws1.max_row 
mc = ws1.max_column 

#RELATIVE SYNONYMOUS CODON USAGE CALCULATION
translation_list = {'*':['TAA','TAG','TGA'],
                    'A':['GCA','GCC','GCG','GCT'],
                    'C':['TGC','TGT'],
                    'D':['GAC','GAT'],
                    'E':['GAA','GAG'],
                    'F':['TTC','TTT'],
                    'G':['GGA','GGC','GGG','GGT'],
                    'H':['CAC','CAT'],
                    'I':['ATA','ATC','ATT'],
                    'K':['AAA','AAG'],
                    'L':['CTA','CTC','CTG','CTT','TTA','TTG'],
                    'M':['ATG'],
                    'N':['AAC','AAT'],
                    'P':['CCA','CCC','CCG','CCT'],
                    'Q':['CAA','CAG'],
                    'R':['AGA','AGG','CGA','CGC','CGG','CGT'],
                    'S':['AGC','AGT','TCA','TCC','TCG','TCT'],
                    'T':['ACA','ACC','ACG','ACT'],
                    'V':['GTA','GTC','GTG','GTT'],
                    'W':['TGG'],
                    'Y':['TAC','TAT']}


# copying the cell values from source  
# excel file to destination excel file 
for i in range (1, mr + 1): 
    codon_dict = {}
    rscu = {}
    for j in range (1, mc + 1): 
        # reading cell value from source excel file 
        c = ws1.cell(row = i, column = j)
        h = ws1.cell(row = 1, column = j)

        c_val = c.value
        h_val = h.value
        
        ws2.cell(row = i, column = j).value = c.value 
        
        #RSCU values are CodonCount/((1/num of synonymous codons) * sum of all synonymous codons)
        if (i>1) and (j>5):
            for aa in translation_list:
                s_codons = translation_list[aa]
                
                for s in s_codons:
                    count = c_val
                    codon_dict[h_val] = count

    #print(codon_dict)
    
    for aa in translation_list:
        total = 0.0
        s_codons = translation_list[aa]

        #THIS WILL COUNT HOW MANY TIMES EACH CODON APPEARS IN THE SEQUENCE
        for s in s_codons:
            count = codon_dict.get(s)
            if type(count) == int:
                total += count

        #calculate the RSCU value for each of the codons
        for  s in s_codons:
            count = codon_dict.get(s)
            if type(count) == int:
                denominator = float(total)/ len(s_codons)
                if denominator == 0:
                    r = 0.0
                else:
                    r = Decimal(count/denominator)
                result = round(r,2)
                rscu[s] = result
        
    final_rscu = {}
    for key in codon_dict:
        value = rscu.get(key)
        final_rscu[key] = value

    for j in range (5, mc + 1):
        if (i != 1):
            # reading cell value from source excel file 
            c = ws1.cell(row = i, column = j)
            h = ws1.cell(row = 1, column = j)

            c_val = c.value
            h_val = h.value
        
            ws2.cell(row = i, column = j).value = final_rscu.get(h_val) 


wb.save(filename = 'CodonTable.xlsx')